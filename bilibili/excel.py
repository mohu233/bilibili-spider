"""
Excel 读写工具
==============
从 bilibili.py 中拆分出的 Excel 读写模块。

功能：
  - load_uids_from_excel()    — 读取 UID 列表（带 7 天去重）
  - write_back_results()      — 写回爬取结果
  - repair_excel()            — 检查/修复损坏的 .xlsx
"""

import os
import re
import json
import shutil
import tempfile
import traceback
from datetime import datetime, timedelta
from typing import List, Optional, Dict

import openpyxl

from config import settings


def _clean_excel_text(value):
    """清洗字符串中 Excel 不支持的非法字符"""
    if value is None:
        return None
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f​‌‍‎‏﻿]', '', str(value))


def _backup_file(filepath):
    """创建 .bak 备份文件"""
    bak_path = filepath + ".bak"
    try:
        shutil.copy2(filepath, bak_path)
    except Exception:
        pass


def load_uids_from_excel(filepath: str) -> List[dict]:
    """
    从 Excel 文件中读取需要更新的 UID 列表。
    自动跳过 7 天内已更新的行（断点续爬）。

    支持的列名（不区分大小写）：
      - uid / mid / 用户id / up主id → UID 列
      - 昵称 / name → 昵称列（可选）
      - last_updated / 更新时间 → 更新时间列
      - hit / 命中 / 备注 → 命中条件列（可选）

    Returns:
        [{"row": 行号, "uid": UID, "name": 昵称, "last_updated": ..., "hit": ...}, ...]
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # 列名映射
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip().lower()] = i

    # UID 列
    uid_col = None
    for key in ["uid", "mid", "用户id", "up主id", "up主uid", "up_id"]:
        if key in col_map:
            uid_col = col_map[key]
            break
    if uid_col is None:
        for key, idx in col_map.items():
            if "id" in key:
                uid_col = idx
                break
    if uid_col is None:
        raise ValueError("无法在 Excel 中找到 UID 列！")

    # 昵称列 & 更新时间列 & 命中列
    name_col = None
    last_updated_col = None
    hit_col = None
    for key, idx in col_map.items():
        if key in ("昵称", "name", "用户名", "username", "uname"):
            name_col = idx
        if key in ("last_updated", "更新时间", "last_update", "updated_at", "update_time", "更新日期"):
            last_updated_col = idx
        if key in ("hit", "命中", "备注", "命中规则", "爬虫命中", "is_furry", "note"):
            hit_col = idx

    # 读取数据行
    raw_rows = []
    for row_idx in range(2, ws.max_row + 1):
        uid_cell = ws.cell(row=row_idx, column=uid_col + 1).value
        if uid_cell is None:
            continue
        try:
            uid = int(str(uid_cell).strip())
        except (ValueError, TypeError):
            continue

        name = ws.cell(row=row_idx, column=name_col + 1).value if name_col is not None else None
        last_val = ws.cell(row=row_idx, column=last_updated_col + 1).value if last_updated_col is not None else None
        hit_val = ws.cell(row=row_idx, column=hit_col + 1).value if hit_col is not None else None

        raw_rows.append({
            "row": row_idx,
            "uid": uid,
            "name": str(name) if name else None,
            "last_updated": str(last_val).strip() if last_val else None,
            "hit": str(hit_val).strip() if hit_val else None,
        })

    wb.close()

    # 断点续爬：过滤 7 天内已更新的行
    now = datetime.now()
    need_update = []
    skip_count = 0
    for row in raw_rows:
        last_str = row.get("last_updated")
        if last_str:
            try:
                last_dt = None
                for fmt in [
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S",
                ]:
                    try:
                        last_dt = datetime.strptime(last_str[:19], fmt)
                        break
                    except ValueError:
                        continue
                if last_dt is None:
                    last_dt = datetime.fromisoformat(last_str)
                if now - last_dt < timedelta(days=7):
                    skip_count += 1
                    continue
            except (ValueError, TypeError):
                pass
        need_update.append(row)

    if skip_count > 0:
        print(f"  [信息] {skip_count} 个 UID 在 7 天内已更新，跳过")
    print(f"  [信息] 需要更新的 UID: {len(need_update)} 个")
    return need_update


def write_back_results(filepath: str, results: List[dict]) -> int:
    """将爬取结果写回 Excel（原子写入：先写临时文件再覆盖）"""
    _backup_file(filepath)

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(filepath) or ".")
    os.close(tmp_fd)

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        # 列名别名映射
        col_aliases = {
            "avatar_url":  ["avatar_url", "头像", "头像url", "face", "face_url"],
            "fan_count":   ["fan_count", "粉丝数", "follower", "fans", "follower_count"],
            "follow_count":["follow_count", "关注数", "following", "following_count"],
            "intro":       ["intro", "简介", "sign", "description", "desc"],
            "hit":         ["hit", "命中", "备注", "命中规则", "爬虫命中", "is_furry", "note"],
            "update_time": ["update_time", "更新日期", "更新时间", "last_updated", "last_update", "updated_at"],
        }

        def find_or_create_col(col_key: str, default_name: str) -> int:
            lower_headers = [str(h).strip().lower() if h else "" for h in headers]
            for alias in col_aliases.get(col_key, [default_name]):
                if alias in lower_headers:
                    return lower_headers.index(alias)
            col_idx = len(headers)
            ws.cell(row=1, column=col_idx + 1, value=default_name)
            headers.append(default_name)
            return col_idx

        avatar_col = find_or_create_col("avatar_url", "avatar_url")
        follower_col = find_or_create_col("fan_count", "fan_count")
        following_col = find_or_create_col("follow_count", "follow_count")
        intro_col = find_or_create_col("intro", "intro")
        hit_col = find_or_create_col("hit", "hit")
        update_time_col = find_or_create_col("update_time", "update_time")

        today_str = datetime.now().strftime("%Y-%m-%d")

        updated = 0
        for r in results:
            row_idx = r["row"]
            if r.get("avatar_url"):
                ws.cell(row=row_idx, column=avatar_col + 1, value=_clean_excel_text(r["avatar_url"]))
            if r.get("follower") is not None:
                ws.cell(row=row_idx, column=follower_col + 1, value=r["follower"])
            if r.get("following") is not None:
                ws.cell(row=row_idx, column=following_col + 1, value=r["following"])
            if r.get("intro") is not None:
                ws.cell(row=row_idx, column=intro_col + 1, value=_clean_excel_text(r["intro"]))
            if r.get("hit") is not None:
                ws.cell(row=row_idx, column=hit_col + 1, value=_clean_excel_text(r["hit"]))
            if r.get("avatar_url") or r.get("follower") is not None:
                ws.cell(row=row_idx, column=update_time_col + 1, value=today_str)
            updated += 1

        wb.save(tmp_path)
        wb.close()
        shutil.move(tmp_path, filepath)
        return updated
    except Exception:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        raise


def repair_excel(filepath: str) -> bool:
    """检查 Excel 是否损坏，损坏则尝试从 .bak 恢复"""
    import zipfile
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            zf.namelist()
        return True
    except zipfile.BadZipFile:
        print(f"  [修复] 检测到文件损坏，尝试从 .bak 恢复…")
        bak_path = filepath + ".bak"
        if os.path.exists(bak_path):
            try:
                with zipfile.ZipFile(bak_path, "r") as zf:
                    zf.namelist()
                shutil.copy2(bak_path, filepath)
                print(f"  [修复] 已从 {bak_path} 恢复")
                return True
            except Exception:
                print(f"  [修复] 备份文件也损坏了，无法恢复")
                return False
        else:
            print(f"  [修复] 没有找到备份文件 {bak_path}")
            return False


def append_new_uids(filepath: str, new_uids: List[dict]) -> None:
    """
    追加新发现的 UID 到 Excel 中（BFS 新节点用）。

    Args:
        filepath: Excel 文件路径
        new_uids: [{"uid": int, "name": str or None, "source_uid": int, "relation": "follower"|"following"}, ...]
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # 获取列索引
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip().lower()] = i

    uid_col = None
    for key in ["uid", "mid"]:
        if key in col_map:
            uid_col = col_map[key]
            break

    name_col = col_map.get("name")
    hit_col = col_map.get("hit")
    avatar_col = col_map.get("avatar_url")
    intro_col = col_map.get("intro")
    fan_col = col_map.get("fan_count")
    following_col = col_map.get("follow_count")

    # 读取已有 UID 去重
    existing_uids = set()
    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        cell = ws.cell(row=row_idx, column=(uid_col or 1) + 1).value
        if cell is not None:
            try:
                existing_uids.add(int(str(cell).strip()))
            except (ValueError, TypeError):
                pass

    # 追加新行
    new_count = 0
    for item in new_uids:
        uid = item["uid"]
        if uid in existing_uids:
            continue
        existing_uids.add(uid)
        next_row = ws.max_row + 1
        if uid_col is not None:
            ws.cell(row=next_row, column=uid_col + 1, value=uid)
        if name_col is not None and item.get("name"):
            ws.cell(row=next_row, column=name_col + 1, value=item["name"])
        if hit_col is not None and item.get("hit"):
            ws.cell(row=next_row, column=hit_col + 1, value=item["hit"])
        if avatar_col is not None and item.get("avatar_url"):
            ws.cell(row=next_row, column=avatar_col + 1, value=_clean_excel_text(item["avatar_url"]))
        if intro_col is not None and item.get("intro"):
            ws.cell(row=next_row, column=intro_col + 1, value=_clean_excel_text(item["intro"]))
        if fan_col is not None and item.get("follower") is not None:
            ws.cell(row=next_row, column=fan_col + 1, value=item["follower"])
        if following_col is not None and item.get("following") is not None:
            ws.cell(row=next_row, column=following_col + 1, value=item["following"])
        new_count += 1

    if new_count > 0:
        wb.save(filepath)
        print(f"  [信息] Excel 新增 {new_count} 个 UID")
    wb.close()
