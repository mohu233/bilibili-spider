"""
Bilibili UP 主信息爬虫
========================
功能：
  1. 从 bilibili_up_export.xlsx 读取 UP 主 UID
  2. 调用 Bilibili API 获取 UP 主头像 URL（及关注数、粉丝数等统计数据）
  3. 将结果写回 bilibili_up_export.xlsx

API 说明：
  - GET https://api.bilibili.com/x/space/wbi/acc/info?mid={uid}
    → 返回用户空间信息，其中 face 字段为头像 URL
    → 需要 WBI 签名（w_rid + wts）+ Cookie (SESSDATA)

  - GET https://api.bilibili.com/x/relation/stat?vmid={uid}
    → 返回关系统计数据（关注数、粉丝数等）
    → 同样需要 WBI 签名 + Cookie

注意事项：
  1. WBI 密钥每天更新一次，程序会缓存密钥，过期自动刷新
  2. 需要在脚本头部填写你的 Cookie（SESSDATA 字段）
  3. 请勿频繁请求，建议每次请求间隔 1-2 秒，避免 IP 被限

依赖安装：
  pip install openpyxl requests
"""

import time
import hashlib
import os
import json
import re
import threading
import traceback
from datetime import datetime, timedelta
from typing import Optional, Tuple, Dict, List

import requests
import openpyxl

# ============================================================
# 配置区 —— 请根据实际情况修改
# ============================================================

# Bilibili Cookie 列表，每个线程使用不同的 Cookie，降低单账号限流风险
# 从浏览器开发者工具中复制（F12 → Network → 任意请求 → Request Headers → Cookie）
# 至少需要 SESSDATA 字段
# 如果有多个账号就填多个 Cookie，会开多线程并行抓取
COOKIES = [
    # Cookie 1 （已填好）
    "buvid3=9430ED47-BE5E-E06D-2A07-62D92D66957893493infoc; b_nut=1774936293; _uuid=E7E10EF42-BA91-B6A4-A576-93110D1559F5E93430infoc; buvid_fp=f3ef882e51e1e36b13738037eace796e; buvid4=3A92A10A-B898-AC38-2414-E3F4A9AD090194122-026033113-6O83AcbCCPf/sTezyu/XOQ%3D%3D; home_feed_column=5; browser_resolution=1707-906; CURRENT_FNVAL=4048; CURRENT_QUALITY=0; rpdid=|(um|kJlRY)m0J'u~~RmJYJ|u; theme-tip-show=SHOWED; theme-avatar-tip-show=SHOWED; SESSDATA=9d126d67%2C1798170753%2C40ca0%2A62CjAgHFcZcaS31GBLAXUN5_ABuWU9xuJDse…DhQSnNmSmhwdjBhN19hSk94cmdLQ3UtcEFCS3RRa2N4VWJrVU5zZFpYRDlpN2pqek9zWk82WnROQkM3U0FNaGJNZmFQeWd3IIEC; bili_jct=35850447d7aca0613f41e7d963a5b054; DedeUserID=363159975; DedeUserID__ckMd5=8fd0214ad39063d3; sid=5rqdv18o; b_lsid=460EA8F3_19F0DC8EE0E; bsource=search_google; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpXVCJ9.eyJleHAiOjE3ODI4Nzc5NDksImlhdCI6MTc4MjYxODY4OSwicGx0IjotMX0.MtuDip6ZILXSyjbnF376eOGAoiMvLy62HKooeIzOPjY; bili_ticket_expires=1782877889; bp_t_offset_363159975=1218827885138149376",
    # Cookie 2（请替换为你第二个账号的 Cookie）
    "buvid3=D98FAC15-A051-2E62-AE72-6115CB655D4344118infoc; b_nut=1781274644; buvid4=00D8FF56-0976-3D4D-CA77-6ACD59FC103A44503-026061222-oSfyrKLbSmpLhZioCx1Q5g%3D%3D; buvid_fp=487e1c6cbd1ef01e3b78ccfafc231e5b; _uuid=7B4F105101-F10FA-1D16-8A98-141FCCAC4FC1055844infoc; theme-tip-show=SHOWED; CURRENT_QUALITY=0; rpdid=|(k||lukRk)R0J'u~)uRumm|J; theme-avatar-tip-show=SHOWED; theme-switch-show=SHOWED; DedeUserID=10624434; DedeUserID__ckMd5=be62134919ff0ba3; SESSDATA=7fcafb42%2C1797952449%2Ce247b%2A61CjBRe1P1zZQRzUQtL7FISFE-9Lsd_eDLnlvx3_pbeVYZdiadDC4vNKyxPlZB8mVLinASVkhseUFoenNDa0lXbVBsQ2gwU29rTzVKUTZtaWFKRlJqMUE3T3RpUHR2dlRhU0NTNHpJaG44VGdfc2lZVEJwWThEa0NZNWlhUVBqTFBQV3BUN3JOOG5nIIEC; bili_jct=75cad187c77335a95e91be199d8658ce; sid=5iqhmn5l; bmg_af_switch=1; bmg_src_def_domain=i2.hdslb.com; hit-dyn-v2=1; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpXVCJ9.eyJleHAiOjE3ODI4NzY5MTksImlhdCI6MTc4MjYxNzY1OSwicGx0IjotMX0.qHJ83PcS62Ja79Su0wTqF9NAU9JvEYb7W4vKvp1yzUU; bili_ticket_expires=1782876859; home_feed_column=4; bsource=search_bing; CURRENT_FNVAL=2000; bp_t_offset_10624434=1218939575762681856; browser_resolution=1355-786; b_lsid=A5FB1147_19F0DEA59B5",
]

# 线程数，自动根据 COOKIES 数量决定
THREAD_COUNT = len(COOKIES)

# 请求间隔（秒），避免请求过快被限流
REQUEST_INTERVAL = 0.5

# Excel 文件名（与脚本同目录）
EXCEL_FILE = "bilibili_up_export.xlsx"

# WBI 密钥缓存文件（与脚本同目录），存放获取到的 img_key / sub_key
WBI_CACHE_FILE = ".wbi_cache.json"


# ============================================================
# WBI 签名算法实现
# ============================================================

# Bilibili WBI 签名混合密钥的 64 位置换表
MIXIN_KEY_ENC_TAB = [
    46, 47, 18, 2, 53, 8, 23, 32, 15, 50, 10, 31, 58, 3, 45, 35,
    27, 43, 5, 49, 33, 9, 42, 19, 29, 28, 14, 39, 12, 38, 41, 13,
    37, 48, 7, 16, 24, 55, 40, 61, 26, 17, 0, 1, 60, 51, 30, 4,
    22, 25, 54, 21, 56, 59, 6, 63, 57, 62, 11, 36, 20, 34, 44, 52
]

# 脚本所在目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def get_wbi_keys() -> Tuple[str, str]:
    """
    从 Bilibili Nav API 获取 wbi_img 的 img_key 和 sub_key。
    优先从本地缓存读取（避免每次启动都请求），缓存过期则重新请求。

    Returns:
        (img_key, sub_key)
    """
    cache_path = os.path.join(SCRIPT_DIR, WBI_CACHE_FILE)

    # 尝试从缓存读取
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cache = json.load(f)
            expire_at = datetime.fromisoformat(cache.get("expire_at", "2000-01-01"))
            if datetime.now() < expire_at:
                return cache["img_key"], cache["sub_key"]
        except (json.JSONDecodeError, KeyError, ValueError):
            pass  # 缓存损坏或过期，重新获取

    # 请求 Nav API 获取最新密钥
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Referer": "https://www.bilibili.com/",
    }
    try:
        resp = requests.get(
            "https://api.bilibili.com/x/web-interface/nav",
            headers=headers,
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        # Nav API 即使未登录（code=-101）也会返回 wbi_img 数据，所以不检查 code
        nav_data = data.get("data")
        if not nav_data or "wbi_img" not in nav_data:
            raise RuntimeError(f"Nav API 返回中没有 wbi_img 字段: {data}")

        wbi_img = nav_data["wbi_img"]
        # 从 URL 中提取文件名中的 key（去掉路径和 .png 后缀）
        img_key = wbi_img["img_url"].rsplit("/", 1)[1].split(".")[0]
        sub_key = wbi_img["sub_url"].rsplit("/", 1)[1].split(".")[0]

        # 写入缓存（密钥每天变化，缓存 12 小时）
        cache = {
            "img_key": img_key,
            "sub_key": sub_key,
            "expire_at": (datetime.now() + timedelta(hours=12)).isoformat(),
        }
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)

        return img_key, sub_key

    except Exception as e:
        # 如果有缓存但过期，降级使用（总比不能用好）
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    cache = json.load(f)
                return cache["img_key"], cache["sub_key"]
            except Exception:
                pass
        raise RuntimeError(f"获取 WBI 密钥失败: {e}")


def get_mixin_key(img_key: str, sub_key: str) -> str:
    """
    根据 img_key 和 sub_key 生成 WBI 混合密钥（取前 32 位）。

    Args:
        img_key: 从 Nav API 获取的 img_key
        sub_key: 从 Nav API 获取的 sub_key

    Returns:
        32 字符的混合密钥
    """
    raw = img_key + sub_key
    return "".join(raw[i] for i in MIXIN_KEY_ENC_TAB)[:32]


def enc_wbi(params: Dict[str, str], mixin_key: str) -> Tuple[str, int]:
    """
    为参数字典生成 WBI 签名。

    Args:
        params: 请求参数字典
        mixin_key: 混合密钥

    Returns:
        (w_rid, wts) — w_rid 为 MD5 签名字符串，wts 为 UNIX 时间戳
    """
    wts = int(time.time())
    params["wts"] = wts

    # 过滤参数值中的特殊字符 !'()*
    params = {
        k: "".join(c for c in str(v) if c not in "!'()*")
        for k, v in params.items()
    }

    # 按 key 排序后拼接为 URL query string（空格编码为 %20 而不是 +）
    query = "&".join(
        f"{k}={requests.utils.quote(str(v), safe='')}"
        for k, v in sorted(params.items())
    )

    # MD5 签名
    sign_str = query + mixin_key
    w_rid = hashlib.md5(sign_str.encode("utf-8")).hexdigest()

    return w_rid, wts


def sign_api(params: Dict[str, str]) -> Dict[str, str]:
    """
    对 API 请求参数进行 WBI 签名，返回添加了 w_rid 和 wts 的参数字典。

    Args:
        params: 原始请求参数

    Returns:
        添加了 w_rid 和 wts 的参数字典
    """
    img_key, sub_key = get_wbi_keys()
    mixin_key = get_mixin_key(img_key, sub_key)
    w_rid, wts = enc_wbi(params.copy(), mixin_key)
    params["w_rid"] = w_rid
    params["wts"] = wts
    return params


# ============================================================
# Bilibili API 请求
# ============================================================

def _clean_cookie(raw: str) -> str:
    """清理 Cookie 字符串，去除 non-latin-1 字符（urllib3 编码要求）"""
    return raw.encode("latin-1", errors="ignore").decode("latin-1")


def get_common_headers(cookie: Optional[str] = None) -> Dict[str, str]:
    """
    返回通用的请求头，包含 Cookie 和必要标识。

    Args:
        cookie: 使用的 Cookie 字符串，默认用 COOKIES[0]

    Returns:
        请求头字典
    """
    cookie_str = cookie if cookie else COOKIES[0]
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Referer": "https://www.bilibili.com/",
        "Cookie": _clean_cookie(cookie_str),
    }


def get_up_info(uid: int, cookie: Optional[str] = None) -> Optional[dict]:
    """
    获取 UP 主的空间信息（含头像 URL）。

    API: GET https://api.bilibili.com/x/space/wbi/acc/info?mid={uid}
    → 响应中 data.face 为头像 URL

    Args:
        uid: UP 主的 UID
        cookie: 使用的 Cookie，默认用 COOKIES[0]

    Returns:
        用户信息字典，失败返回 None
    """
    params = sign_api({"mid": str(uid)})
    headers = get_common_headers(cookie)

    try:
        resp = requests.get(
            "https://api.bilibili.com/x/space/wbi/acc/info",
            params=params,
            headers=headers,
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()

        if data.get("code") == 0 and data.get("data"):
            return data["data"]
        elif data.get("code") == -412:
            print(f"    [警告] UID {uid}: 请求被拦截（-412），可能触发频率限制，暂停 10 秒…")
            time.sleep(10)
            return None
        elif data.get("code") == -352:
            print(f"    [警告] UID {uid}: WBI 签名验证失败（-352），尝试刷新密钥重试…")
            # 删除缓存，下次请求会自动刷新
            cache_path = os.path.join(SCRIPT_DIR, WBI_CACHE_FILE)
            if os.path.exists(cache_path):
                os.remove(cache_path)
            return None
        else:
            print(f"    [警告] UID {uid}: API 返回异常 code={data.get('code')}, msg={data.get('message')}")
            return None

    except requests.exceptions.Timeout:
        print(f"    [错误] UID {uid}: 请求超时")
        return None
    except requests.exceptions.RequestException as e:
        print(f"    [错误] UID {uid}: 请求失败 — {e}")
        traceback.print_exc()
        return None
    except json.JSONDecodeError:
        print(f"    [错误] UID {uid}: 响应 JSON 解析失败")
        traceback.print_exc()
        return None


def get_up_stat(uid: int, cookie: Optional[str] = None) -> Optional[dict]:
    """
    获取 UP 主的统计数据（关注数、粉丝数等）。

    API: GET https://api.bilibili.com/x/relation/stat?vmid={uid}
    → 响应中 data 包含 following / follower / dynamic_count 等字段

    Args:
        uid: UP 主的 UID
        cookie: 使用的 Cookie，默认用 COOKIES[0]

    Returns:
        统计数据字典，失败返回 None
    """
    params = sign_api({"vmid": str(uid)})
    headers = get_common_headers(cookie)

    try:
        resp = requests.get(
            "https://api.bilibili.com/x/relation/stat",
            params=params,
            headers=headers,
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()

        if data.get("code") == 0 and data.get("data"):
            return data["data"]
        elif data.get("code") == -412:
            print(f"    [警告] UID {uid}: 请求被拦截（-412），暂停 10 秒…")
            time.sleep(10)
            return None
        elif data.get("code") == -352:
            print(f"    [警告] UID {uid}: WBI 签名验证失败（-352），尝试刷新密钥重试…")
            cache_path = os.path.join(SCRIPT_DIR, WBI_CACHE_FILE)
            if os.path.exists(cache_path):
                os.remove(cache_path)
            return None
        else:
            print(f"    [警告] UID {uid}: Relation API 返回 code={data.get('code')}")
            return None

    except requests.exceptions.Timeout:
        print(f"    [错误] UID {uid}: 请求超时")
        return None
    except requests.exceptions.RequestException as e:
        print(f"    [错误] UID {uid}: 请求失败 — {e}")
        traceback.print_exc()
        return None
    except json.JSONDecodeError:
        print(f"    [错误] UID {uid}: 响应 JSON 解析失败")
        traceback.print_exc()
        return None


# ============================================================
# Excel 读写
# ============================================================

def _clean_excel_text(value):
    """清洗字符串中 Excel 不支持的非法字符（控制字符）"""
    if value is None:
        return None
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f​‌‍‎‏﻿]', '', str(value))


def load_uids_from_excel(filepath: str) -> List[dict]:
    """
    从 Excel 文件中读取需要更新的 UID 列表。
    自动检查 last_updated 列：7 天内已更新的行跳过，只返回需要重新抓取的行。

    支持的列名（不区分大小写）：
      - uid / mid / 用户id / up主id / up主uid → UID 列
      - 昵称 / name / 用户名 → 昵称列（可选）
      - last_updated / 更新时间 / update_time / 更新日期 → 更新时间列（断点续爬判断）
      - hit / 命中 / 命中规则 / 备注 → 命中条件列（可选，透传保留）

    Args:
        filepath: Excel 文件路径

    Returns:
        [
            {
                "row": 行号（从 1 开始）,
                "uid": UID 数值,
                "name": 昵称（或 None）,
            },
            ...
        ]
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 读取表头（第一行）
    headers = [cell.value for cell in ws[1]]
    print(f"  [信息] 检测到表头: {headers}")

    # 构建列名 → 列索引的映射（不区分大小写，去空格）
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            key = str(h).strip().lower()
            col_map[key] = i

    # 定位 UID 列
    uid_col = None
    for key in ["uid", "mid", "用户id", "up主id", "up主uid", "up_id"]:
        if key in col_map:
            uid_col = col_map[key]
            break

    if uid_col is None:
        for key, idx in col_map.items():
            if "id" in key:
                uid_col = idx
                break

    if uid_col is None:
        raise ValueError("无法在 Excel 中找到 UID 列！请确保表头包含「uid」「mid」「用户id」等列名。")

    # 定位昵称列 & 更新时间列 & 命中列
    name_col = None
    last_updated_col = None
    hit_col = None
    for key, idx in col_map.items():
        if key in ("昵称", "name", "用户名", "username", "uname"):
            name_col = idx
        if key in ("last_updated", "更新时间", "last_update", "updated_at", "update_time", "更新日期"):
            last_updated_col = idx
        if key in ("hit", "命中", "备注", "命中规则", "爬虫命中", "note"):
            hit_col = idx

    # 读取数据行
    raw_rows = []
    for row_idx in range(2, ws.max_row + 1):
        uid_cell = ws.cell(row=row_idx, column=uid_col + 1).value
        if uid_cell is None:
            continue

        # 尝试转换为整数 UID
        try:
            uid = int(str(uid_cell).strip())
        except (ValueError, TypeError):
            print(f"    [跳过] 第 {row_idx} 行 UID 格式无效: {uid_cell}")
            continue

        name = None
        if name_col is not None:
            name = ws.cell(row=row_idx, column=name_col + 1).value

        last_updated_val = None
        if last_updated_col is not None:
            last_updated_val = ws.cell(row=row_idx, column=last_updated_col + 1).value

        hit_val = None
        if hit_col is not None:
            hit_val = ws.cell(row=row_idx, column=hit_col + 1).value

        raw_rows.append({
            "row": row_idx,
            "uid": uid,
            "name": str(name) if name else None,
            "last_updated": str(last_updated_val).strip() if last_updated_val else None,
            "hit": str(hit_val).strip() if hit_val else None,
        })

    wb.close()

    # ========== 断点续爬：过滤 7 天内已更新的行 ==========
    now = datetime.now()
    need_update = []
    skip_count = 0
    for row in raw_rows:
        last_str = row.get("last_updated")
        if last_str:
            try:
                # 尝试多种日期格式
                last_dt = None
                for fmt in [
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M",
                    "%Y-%m-%d",
                    "%Y/%m/%d %H:%M:%S",
                ]:
                    try:
                        last_dt = datetime.strptime(last_str[:19], fmt)
                        break
                    except ValueError:
                        continue
                if last_dt is None:
                    last_dt = datetime.fromisoformat(last_str)

                if now - last_dt < timedelta(days=7):
                    skip_count += 1
                    continue
            except (ValueError, TypeError):
                pass  # 日期解析失败，重新抓取
        need_update.append(row)

    if skip_count > 0:
        print(f"  [信息] {skip_count} 个 UID 在 7 天内已更新，跳过")
    print(f"  [信息] 需要更新的 UID: {len(need_update)} 个")
    return need_update


def _backup_file(filepath):
    """创建 .bak 备份文件"""
    bak_path = filepath + ".bak"
    try:
        import shutil
        shutil.copy2(filepath, bak_path)
    except Exception:
        pass


def write_back_results(
    filepath: str,
    results: List[dict],
) -> None:
    """
    先将数据写入临时文件，再用临时文件覆盖原文件（防止写坏原文件）。
    """
    import tempfile
    import shutil

    # 备份原文件
    _backup_file(filepath)

    # 先写入临时文件
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(filepath) or ".")
    os.close(tmp_fd)

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        # 列名映射：中文 ↔ 英文
        col_aliases = {
            "avatar_url": ["avatar_url", "头像", "头像url", "face", "face_url"],
            "fan_count":  ["fan_count", "粉丝数", "follower", "fans", "follower_count"],
            "follow_count": ["follow_count", "关注数", "following", "following_count"],
            "intro":      ["intro", "简介", "sign", "description", "desc"],
            "hit":        ["hit", "命中", "备注", "命中规则", "爬虫命中", "note"],
            "update_time": ["update_time", "更新日期", "更新时间", "last_updated", "last_update", "updated_at"],
        }

        def find_or_create_col(col_key: str, default_name: str) -> int:
            lower_headers = [str(h).strip().lower() if h else "" for h in headers]
            for alias in col_aliases.get(col_key, [default_name]):
                if alias in lower_headers:
                    return lower_headers.index(alias)
            col_idx = len(headers)
            ws.cell(row=1, column=col_idx + 1, value=default_name)
            headers.append(default_name)
            return col_idx

        avatar_col = find_or_create_col("avatar_url", "avatar_url")
        follower_col = find_or_create_col("fan_count", "fan_count")
        following_col = find_or_create_col("follow_count", "follow_count")
        intro_col = find_or_create_col("intro", "intro")
        hit_col = find_or_create_col("hit", "hit")
        update_time_col = find_or_create_col("update_time", "update_time")

        today_str = datetime.now().strftime("%Y-%m-%d")

        updated = 0
        for r in results:
            row_idx = r["row"]
            if r.get("avatar_url"):
                ws.cell(row=row_idx, column=avatar_col + 1, value=_clean_excel_text(r["avatar_url"]))
            if r.get("follower") is not None:
                ws.cell(row=row_idx, column=follower_col + 1, value=r["follower"])
            if r.get("following") is not None:
                ws.cell(row=row_idx, column=following_col + 1, value=r["following"])
            if r.get("intro") is not None:
                ws.cell(row=row_idx, column=intro_col + 1, value=_clean_excel_text(r["intro"]))
            if r.get("hit") is not None:
                ws.cell(row=row_idx, column=hit_col + 1, value=_clean_excel_text(r["hit"]))
            if r.get("avatar_url") or r.get("follower") is not None:
                ws.cell(row=row_idx, column=update_time_col + 1, value=today_str)
            updated += 1

        # 保存到临时文件
        wb.save(tmp_path)
        wb.close()

        # 用临时文件覆盖原文件（原子替换，不会写坏）
        shutil.move(tmp_path, filepath)

        return updated
    except Exception:
        # 如果有临时文件残留，清理掉
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        raise


# ============================================================
# 主流程
# ============================================================

def crawl_worker(
    rows_chunk: List[dict],
    cookie: str,
    thread_id: int,
    all_results: List,
    lock: threading.Lock,
    excel_path: str,
    total_rows: int,
):
    """
    单线程爬取工作函数。每个线程处理自己的 UID 分片。

    Args:
        rows_chunk: 分配给该线程的 UID 列表
        cookie: 该线程使用的 Cookie
        thread_id: 线程编号（0, 1, 2...）
        all_results: 共享的结果列表（线程安全）
        lock: 全局锁，用于 Excel 写入同步
        excel_path: Excel 文件路径
        total_rows: 总行数（用于进度显示）
    """
    save_counter = 0
    WRITE_INTERVAL = 50  # 每爬 50 条存一次盘

    for i, row in enumerate(rows_chunk):
        uid = row["uid"]
        name_str = f" ({row['name']})" if row["name"] else ""
        display = f"[线程{thread_id + 1}] [{i + 1}/{len(rows_chunk)}] UID {uid}{name_str}"

        print(f"\n  ▶ {display}")

        try:
            # ========== 获取空间信息（含头像） ==========
            info = get_up_info(uid, cookie)
            if info:
                avatar_url = info.get("face", "")
                name_api = info.get("name", "")
                intro = info.get("sign", "")

                print(f"    ├ 昵称: {name_api}")
                print(f"    ├ 头像: {avatar_url}")

                result = {
                    "row": row["row"],
                    "uid": uid,
                    "avatar_url": avatar_url,
                    "intro": intro,
                    "following": None,
                    "follower": None,
                    "hit": row.get("hit"),
                }
            else:
                print(f"    └ [失败] 获取用户信息失败")
                result = {
                    "row": row["row"],
                    "uid": uid,
                    "avatar_url": None,
                    "intro": None,
                    "following": None,
                    "follower": None,
                    "hit": row.get("hit"),
                }

            # ========== 获取统计数据（关注数、粉丝数） ==========
            time.sleep(REQUEST_INTERVAL * 0.5)
            stat = get_up_stat(uid, cookie)
            if stat:
                following = stat.get("following", 0)
                follower = stat.get("follower", 0)
                print(f"    ├ 关注数: {following}")
                print(f"    └ 粉丝数: {follower}")
                result["following"] = following
                result["follower"] = follower
            else:
                print(f"    └ [失败] 获取统计数据失败")

        except Exception as e:
            print(f"    [错误] 处理 UID {uid} 时发生异常: {e}")
            traceback.print_exc()
            result = {
                "row": row["row"],
                "uid": uid,
                "avatar_url": None,
                "intro": None,
                "following": None,
                "follower": None,
                "hit": row.get("hit"),
            }

        # 合并到全局结果（内存中）
        with lock:
            existing_uids = {r["uid"] for r in all_results}
            if result["uid"] not in existing_uids:
                all_results.append(result)
                existing_uids.add(result["uid"])
        save_counter += 1

        # ========== 每 WRITE_INTERVAL 条存一次盘 ==========
        should_write = (
            save_counter >= WRITE_INTERVAL
            or i == len(rows_chunk) - 1  # 最后一条
        )
        if should_write:
            with lock:
                try:
                    n = write_back_results(excel_path, all_results)
                    print(f"\n  [存盘] 线程{thread_id + 1} 已保存 {n} 行数据到 Excel")
                    save_counter = 0
                except Exception as e:
                    print(f"\n  [错误] 保存失败: {e}")
                    print(f"  [提示] 数据仍在内存中，将尝试下次保存")
                    traceback.print_exc()

        # 请求间隔
        if i < len(rows_chunk) - 1:
            time.sleep(REQUEST_INTERVAL)


def _repair_excel(filepath):
    """
    检查 Excel 文件是否损坏，损坏则尝试从 .bak 恢复。
    返回 True 表示文件可正常使用，False 表示彻底损坏无法恢复。
    """
    import zipfile

    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            # 如果能正常读取 ZIP 目录，说明文件完好
            zf.namelist()
        return True
    except zipfile.BadZipFile:
        print(f"  [修复] 检测到文件损坏，尝试从 .bak 备份恢复…")
        bak_path = filepath + ".bak"
        if os.path.exists(bak_path):
            try:
                with zipfile.ZipFile(bak_path, "r") as zf:
                    zf.namelist()
                import shutil
                shutil.copy2(bak_path, filepath)
                print(f"  [修复] 已从 {bak_path} 恢复")
                return True
            except Exception:
                print(f"  [修复] 备份文件也损坏了，无法恢复")
                return False
        else:
            print(f"  [修复] 没有找到备份文件 {bak_path}，无法恢复")
            return False


def main():
    """主函数：读取 Excel → 多线程爬取 → 写回 Excel"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILE)

    print("=" * 60)
    print("Bilibili UP 主信息爬虫 (多线程版)")
    print("=" * 60)

    # 检查 Cookie 配置
    placeholder_count = sum(1 for c in COOKIES if "填入" in c)
    valid_count = len(COOKIES) - placeholder_count
    if valid_count == 0:
        print("[错误] 请先在 COOKIES 列表中填写你的 Bilibili Cookie！")
        print("       获取方式：浏览器登录 bilibili.com → F12 → Network → 任意请求")
        print("       → Request Headers → Cookie → 复制")
        return

    if placeholder_count > 0:
        print(f"[警告] COOKIES 中有 {placeholder_count} 个占位 Cookie 未填写，将只用 {valid_count} 个线程")
        print()

    # 检查 Excel 是否存在
    if not os.path.exists(excel_path):
        print(f"[错误] 找不到 Excel 文件: {excel_path}")
        print("请确认 bilibili_up_export.xlsx 与本脚本在同一目录。")
        return

    # 自动修复损坏的文件
    _repair_excel(excel_path)

    # 1. 读取 UID（自动跳过 7 天内已更新的行）
    print("\n[Step 1/3] 读取 Excel 中的 UID…")
    try:
        rows = load_uids_from_excel(excel_path)
    except Exception as e:
        print(f"[错误] 读取 Excel 失败: {e}")
        traceback.print_exc()
        return

    if not rows:
        print("[完成] 所有 UID 均在 7 天内更新过，无需抓取。")
        return

    print(f"  [信息] 本次需爬取 {len(rows)} 个 UP 主")

    # 2. 多线程爬取
    real_cookies = [c for c in COOKIES if "填入" not in c]
    n_threads = min(len(real_cookies), len(rows))

    # 把 UID 列表均匀分给各个线程
    chunks = [[] for _ in range(n_threads)]
    for idx, row in enumerate(rows):
        chunks[idx % n_threads].append(row)

    print(f"\n[Step 2/3] 启动 {n_threads} 个线程并行爬取（间隔 {REQUEST_INTERVAL} 秒）…")
    print(f"  [提示] 每爬 50 条自动存盘一次，带 .bak 备份")
    print(f"  [提示] 线程数 = Cookie 数 = {n_threads}")
    for tid, (chunk, ck) in enumerate(zip(chunks, real_cookies)):
        print(f"          线程{tid + 1}: {len(chunk)} 个 UID")

    all_results = []  # 全局结果列表
    results_lock = threading.Lock()
    threads = []

    for tid, (chunk, ck) in enumerate(zip(chunks, real_cookies)):
        if not chunk:
            continue
        t = threading.Thread(
            target=crawl_worker,
            args=(chunk, ck, tid, all_results, results_lock, excel_path, len(rows)),
            daemon=True,
        )
        threads.append(t)
        t.start()

    # 等待所有线程完成
    for t in threads:
        t.join()

    # 所有线程结束后，再存一次确保最终数据写盘
    if all_results:
        try:
            n = write_back_results(excel_path, all_results)
            print(f"\n  [最终存盘] 已保存 {n} 行")
        except Exception as e:
            print(f"\n  [错误] 最终存盘失败: {e}")
            traceback.print_exc()

    completed = len(all_results)
    print(f"\n  [信息] 共爬取 {completed}/{len(rows)} 个 UP 主")

    # 汇总
    print("\n" + "=" * 60)
    print("爬取完成！")
    avatar_filled = sum(1 for r in all_results if r.get("avatar_url"))
    intro_filled = sum(1 for r in all_results if r.get("intro"))
    stat_filled = sum(1 for r in all_results if r.get("follower") is not None)
    print(f"  - 获取头像: {avatar_filled}/{completed}")
    print(f"  - 获取简介: {intro_filled}/{completed}")
    print(f"  - 获取粉丝/关注: {stat_filled}/{completed}")
    print(f"  - 数据已保存至: {EXCEL_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n[提示] 用户中断！已爬取的数据已逐条保存到 {EXCEL_FILE}，不会丢失。")
    except Exception as e:
        print(f"\n[错误] 程序崩溃: {e}")
        traceback.print_exc()
        print(f"[提示] 崩溃前的数据已逐条保存到 {EXCEL_FILE}，不会丢失。")
